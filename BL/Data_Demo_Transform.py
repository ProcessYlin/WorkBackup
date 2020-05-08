# -*- coding: utf-8 -*-
# @Time    : 2019/11/10 1:24 PM
# @Author  : ylin
# @Email   :
# @File    : pd_toexcel.py
# @Software: PyCharm
from openpyxl import Workbook
from openpyxl import load_workbook
from inspect import isfunction
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.cell.cell import WriteOnlyCell
import json
import  datetime
import resource

def PathJson():
    pass

def JsonInput(path):
    with open(path) as load_f:
        load_dict = json.load(load_f)
        file_sheets = len(load_dict['meta']['sheet_mapping'])
        print(file_sheets)
    # print(load_dict)
    return load_dict, file_sheets


def writeExcel2(file_name, load_dict, file_sheets, row_num=2):
    wb = load_workbook(filename=file_name, keep_vba=False)
    sheet_name = wb.sheetnames
    print(datetime.datetime.now())
    # print(len(sheet_name))
    for num in range(len(sheet_name)):
        # 判断sheet位置小于Json数据数量则写入，大于则标示数据插入完毕
        if num <= file_sheets:
            sheet = wb.worksheets[num]
            # print(sheet)
            for data_num, data_json in load_dict['data'].items():
                # 获取数据所属sheet位置
                sheet_num = data_num.split('_')[0]
                history_data =[]
                # 判断Json所属sheet与打开sheet位置是否相同
                if int(sheet_num)==int(num):
                    for m in range(len(load_dict['data'][data_num]['data'])):
                        history_data.append(list(load_dict['data'][data_num]['data'][m].values()))
                    # print(history_data)
                for i in range(len(history_data)):
                    for j in range(len(history_data[i])):
                        # row_num 数据从指定行数开始追加，默认从1
                        sheet.cell(row=i+row_num, column=j+1).value = history_data[i][j]
        else:
            print("数据插入完毕")
            break

    wb.save("/Users/lin/Desktop/New_demo01.xlsx")


def main():
    # 设置该脚本可使用最大内存8G
    try:
        soft, hard = resource.getrlimit(resource.RLIMIT_AS)
        resource.setrlimit(resource.RLIMIT_AS, (9223372036854775807, hard))
        load_dict, file_sheets = JsonInput(path)
        writeExcel2(file_name, load_dict, file_sheets)
    except Exception as e:
        print(e)



if __name__ == '__main__':
    startime = datetime.datetime.now()
    print(startime)
    file_name = "/Users/lin/Desktop/大文件/费用提报(主线)_分区财务.xlsx"
    path = "/Users/lin/Desktop/大文件/BL.json"
    main()
    endtime = datetime.datetime.now()
    print(endtime-startime)
