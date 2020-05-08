# -*- coding: utf-8 -*-
# @Time    : 2019/11/10 1:24 PM
# @Author  : ylin
# @Email   :
# @File    : pd_toexcel.py
# @Software: PyCharm
from openpyxl import Workbook
from openpyxl import load_workbook
from inspect import isfunction
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.cell.cell import WriteOnlyCell
import json
import datetime
import resource
import numpy as np

def writeExcel_brand(file_path,data):
    # print(data.shape)
    # 品牌层面包含的关键字
    area_list = ['华南','华南','东北','华东','华中','鲁豫','西南','西北','总公司','电商']
    [h, l] = data.shape
    wb = load_workbook(filename=file_path, keep_vba=False)
    sheet_name = wb.sheetnames
    print(sheet_name)
    for num in range(len(sheet_name)):
        if any(area in sheet_name[num] for area in area_list) and  ('分级' not in sheet_name[num]):
            sheet = wb.worksheets[num]
            print(sheet_name[num])
            for i in range(h):
                for j in range(l):
                    # row_num 数据从指定行数开始追加，默认从1
                    sheet.cell(row=i+5, column=j+64).value = data[i][j]
    # wb.save("/Users/lin/Desktop/New_demo01.xlsx")
    wb.save(file_path)
    wb.close()


def writeExcel_area(file_path, data):
    # print(data.shape)
    [h, l] = data.shape
    wb = load_workbook(filename=file_path, keep_vba=False)
    sheet_name = wb.sheetnames
    print(sheet_name)
    for num in range(len(sheet_name)):
        if ('分区'  in sheet_name[num]) and ('分级' not in sheet_name[num]):
            sheet = wb.worksheets[num]
            print(sheet_name[num])
            for i in range(h):
                for j in range(l):
                    # row_num 数据从指定行数开始追加，默认从1
                    sheet.cell(row=i+6, column=j+64).value = data[i][j]
    # wb.save("/Users/lin/Desktop/New_demo01.xlsx")
    wb.save(file_path)
    wb.close()


def main():
    # 填充的主数据
    data = np.array([
        [0 		,1158 	,0 		,202 	,0 		,1	    ,0 	    ,0   ,0    ,0],
        [0 		,933 	,0 		,169 	,0 		,1	    ,0 	    ,0   ,0    ,0],
        [0 		,831 	,0 		,130 	,0 		,1	    ,0 	    ,0   ,0    ,0],
        [0 		,933 	,0 		,163 	,0 		,1	    ,0 	    ,141 ,0    ,0],
        [0 		,1171 	,0 		,188 	,0 		,1	    ,0 	    ,166 ,0    ,0],
        [0 		,896 	,0 		,156 	,0 		,1	    ,0 	    ,229 ,0    ,0],
        [0 		,837 	,0 		,130 	,0 		,1	    ,0 	    ,265 ,0    ,0],
        [0 		,924 	,0 		,156 	,0 		,0.97	,0 	    ,311 ,0    ,0],
        [0 		,1187 	,0 		,189 	,0 		,0.85	,0 	    ,340 ,500  ,0],
        [880 	,880 	,150 	,150 	,0.67   ,0 	    ,404 	,322 ,468  ,0],
        [830 	,830 	,128 	,128 	,0.69   ,0 	    ,415 	,332 ,5000 ,0],
        [920 	,0 		,155 	,0 		,0.68   ,0 	    ,443 	,0   ,5000 ,0],
        [1170 	,0 		,190 	,0 		,0.737  ,0 	    ,565 	,0   ,0    ,0],
        [940 	,0 		,175 	,0 		,0      ,0 	    ,511 	,0   ,0    ,0],
        [930 	,0 		,130 	,0 		,0      ,0 	    ,530 	,0   ,0    ,0]
        ])
    # 设置该脚本可使用最大内存8G
    try:
        soft, hard = resource.getrlimit(resource.RLIMIT_AS)
        resource.setrlimit(resource.RLIMIT_AS, (9223372036854775807, hard))
        if level == 1:
            writeExcel_brand(file_name,data)
        elif level == 2:
            writeExcel_area(file_name,data)
        else:
            print('未知的级别')
    except Exception as e:
        print(e)



if __name__ == '__main__':
    startime = datetime.datetime.now()
    file_name = "/Users/lin/Desktop/商品预算模版-拔佳品牌的副本.xlsx"
    level = 1
    main()
    endtime = datetime.datetime.now()
    print(endtime-startime)